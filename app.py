import os
import io
import logging
from flask import Flask, request, render_template, send_file, flash, jsonify
from werkzeug.utils import secure_filename
from werkzeug.middleware.proxy_fix import ProxyFix
import msoffcrypto
from openpyxl import load_workbook
from openpyxl.workbook.protection import WorkbookProtection

# Configure logging
logging.basicConfig(level=logging.DEBUG)

# Create the app
app = Flask(__name__)
app.secret_key = os.environ.get("SESSION_SECRET", "dev-secret-key-change-in-production")
app.wsgi_app = ProxyFix(app.wsgi_app, x_proto=1, x_host=1)

# Configuration
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50MB max file size

def allowed_file(filename):
    """Check if the uploaded file has an allowed extension."""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in {'xlsx'}

@app.route('/')
def index():
    """Render the main page with the upload form."""
    return render_template('index.html')

@app.route('/decrypt_excel', methods=['POST'])
def decrypt_excel():
    """
    Decrypt a password-protected Excel file and return it as a downloadable response.
    
    Expected form data:
    - file: The encrypted Excel file (.xlsx)
    - password: The password to decrypt the file
    """
    try:
        # Check if the request contains a file
        if 'file' not in request.files:
            app.logger.error("No file provided in the request")
            return jsonify({'error': 'No file provided'}), 400
        
        file = request.files['file']
        password = request.form.get('password', '')
        
        # Check if a file was actually selected
        if file.filename == '':
            app.logger.error("No file selected")
            return jsonify({'error': 'No file selected'}), 400
        
        # Check if password was provided
        if not password:
            app.logger.error("No password provided")
            return jsonify({'error': 'Password is required'}), 400
        
        # Validate file type
        if not allowed_file(file.filename):
            app.logger.error(f"Invalid file type: {file.filename}")
            return jsonify({'error': 'Only .xlsx files are allowed'}), 400
        
        # Read the uploaded file into memory
        file_data = file.read()
        app.logger.info(f"Read file: {file.filename}, size: {len(file_data)} bytes")
        
        # Create input stream from file data
        input_stream = io.BytesIO(file_data)
        
        # Create output stream for decrypted data
        output_stream = io.BytesIO()
        
        # Log the first few bytes of original file for debugging
        app.logger.debug(f"Original file first 16 bytes: {file_data[:16].hex()}")
        
        try:
            # Initialize msoffcrypto with the encrypted file
            office_file = msoffcrypto.OfficeFile(input_stream)
            
            # Load the password
            office_file.load_key(password=password)
            
            # Decrypt the file to the output stream
            office_file.decrypt(output_stream)
            
            # Get the size of decrypted data for verification
            decrypted_size = output_stream.tell()
            app.logger.info(f"File decrypted successfully. Decrypted size: {decrypted_size} bytes")
            
            # Verify we have actual decrypted content
            if decrypted_size == 0:
                app.logger.error("Decryption resulted in empty file")
                return jsonify({'error': 'Decryption resulted in empty file. Please check your password.'}), 500
            
        except msoffcrypto.exceptions.InvalidKeyError:
            app.logger.error("Invalid password provided")
            return jsonify({
                'error': 'Invalid password. The password you entered is incorrect. Please verify the password and try again.',
                'error_type': 'invalid_password'
            }), 401
        
        except msoffcrypto.exceptions.FileFormatError:
            app.logger.error("Invalid file format or corrupted file")
            return jsonify({
                'error': 'Invalid file format or corrupted file. Please ensure the file is a valid encrypted Excel (.xlsx) file.',
                'error_type': 'invalid_format'
            }), 400
            
        except msoffcrypto.exceptions.DecryptionError:
            app.logger.error("Decryption failed - possibly wrong password")
            return jsonify({
                'error': 'Decryption failed. This usually means the password is incorrect or the file is corrupted.',
                'error_type': 'decryption_failed'
            }), 401
        
        except Exception as e:
            app.logger.error(f"Decryption failed: {str(e)}")
            # Check if it's likely a password issue
            if 'password' in str(e).lower() or 'key' in str(e).lower():
                return jsonify({
                    'error': 'Decryption failed. Please check your password and try again.',
                    'error_type': 'password_related'
                }), 401
            else:
                return jsonify({
                    'error': f'Decryption failed: {str(e)}',
                    'error_type': 'general_error'
                }), 500
        
        # Reset stream position for reading
        output_stream.seek(0)
        
        # Generate filename for the decrypted file
        if file.filename:
            original_filename = secure_filename(file.filename)
            # Remove .xlsx extension and add unprotected suffix
            base_name = original_filename.rsplit('.', 1)[0] if '.' in original_filename else original_filename
            decrypted_filename = f"{base_name}_unprotected.xlsx"
        else:
            decrypted_filename = "unprotected_file.xlsx"
        
        app.logger.info(f"Sending decrypted file: {decrypted_filename}, size: {decrypted_size} bytes")
        
        # Get the decrypted data and log first few bytes for verification
        decrypted_data = output_stream.getvalue()
        app.logger.debug(f"Decrypted file first 16 bytes: {decrypted_data[:16].hex()}")
        
        # Verify the file starts with Excel magic bytes (PK signature for ZIP format)
        if decrypted_data[:2] == b'PK':
            app.logger.info("Decrypted file has correct Excel format signature")
        else:
            app.logger.warning(f"Decrypted file may not be valid Excel format. First 4 bytes: {decrypted_data[:4].hex()}")
        
        # Remove additional Excel protections (worksheet protection, workbook protection)
        try:
            app.logger.info("Removing worksheet and workbook protections...")
            
            # Load the decrypted workbook
            decrypted_stream = io.BytesIO(decrypted_data)
            workbook = load_workbook(decrypted_stream)
            
            # Remove workbook protection
            if workbook.security:
                app.logger.info("Removing workbook protection")
                workbook.security = WorkbookProtection()
            
            # Remove worksheet protections
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                if worksheet.protection.sheet:
                    app.logger.info(f"Removing protection from worksheet: {sheet_name}")
                    worksheet.protection.sheet = False
                    worksheet.protection.password = None
                    worksheet.protection.enable()
            
            # Save the fully unprotected workbook to a new stream
            final_output = io.BytesIO()
            workbook.save(final_output)
            final_output.seek(0)
            
            # Log the size of the final processed file
            final_size = len(final_output.getvalue())
            app.logger.info(f"Successfully removed all Excel protections. Final file size: {final_size} bytes")
            final_output.seek(0)  # Reset position after getvalue()
            
        except Exception as e:
            app.logger.warning(f"Could not remove Excel protections: {str(e)}. Returning decrypted file as-is.")
            # If protection removal fails, return the decrypted file as-is
            final_output = io.BytesIO(decrypted_data)
            app.logger.info(f"Fallback: using decrypted file size: {len(decrypted_data)} bytes")
        
        # Log final details before sending
        final_output.seek(0, 2)  # Seek to end
        final_file_size = final_output.tell()
        final_output.seek(0)  # Reset to beginning
        
        app.logger.info(f"About to send file: {decrypted_filename}")
        app.logger.info(f"Final response file size: {final_file_size} bytes")
        app.logger.info(f"Original encrypted size: {len(file_data)} bytes")
        app.logger.info(f"Size difference: {final_file_size - len(file_data)} bytes")
        
        # Return the fully unprotected Excel file as a downloadable response
        response = send_file(
            final_output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=decrypted_filename
        )
        
        # Add headers to prevent caching
        response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'
        
        return response
        
    except Exception as e:
        app.logger.error(f"Unexpected error: {str(e)}")
        return jsonify({'error': f'An unexpected error occurred: {str(e)}'}), 500

@app.errorhandler(413)
def too_large(e):
    """Handle file too large error."""
    return jsonify({'error': 'File too large. Maximum size allowed is 50MB.'}), 413

@app.errorhandler(404)
def not_found(e):
    """Handle 404 errors."""
    return jsonify({'error': 'Endpoint not found'}), 404

@app.errorhandler(405)
def method_not_allowed(e):
    """Handle method not allowed errors."""
    return jsonify({'error': 'Method not allowed. Use POST for /decrypt_excel endpoint.'}), 405

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)
